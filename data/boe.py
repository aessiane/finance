#!/usr/bin/env python3
#
# Copyright (c) 2024 LateGenXer
#
# SPDX-License-Identifier: AGPL-3.0-or-later
#


import math
import os
import posixpath
import zipfile

import openpyxl

import pandas as pd

from download import download


def read(sh, name):
    # Find the last row
    row = sh.max_row
    for row in range(6, sh.max_row):
        if sh.cell(row + 1, 1).value is None:
            break

    date = sh.cell(row, 1).value.date()

    years_row = 4

    assert sh.cell(years_row, 1).value == 'years:'

    col = 2
    data = []
    for col in range(2, sh.max_column):
        years = sh.cell(years_row,   col).value
        if years is None:
            break
        assert years - math.floor(years) in (0.0, 0.5)
        rate = sh.cell(row, col).value
        data.append((years, rate))

    df = pd.DataFrame(data, columns=['Years', name])
    df.set_index('Years', inplace=True)

    assert df.index.is_monotonic_increasing
    assert df.index.is_unique

    return date, df


_data_dir = os.path.dirname(__file__)
_filename = os.path.join(_data_dir, 'boe-yield-curves.csv')


_measures = 'Nominal', 'Real', 'Inflation'


def load():
    url = 'https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip'
    filename = os.path.join(_data_dir, posixpath.basename(url))
    download(url, filename, content_type='application/x-zip-compressed', ttl=6*3600)
    archive = zipfile.ZipFile(filename)

    dfs = []

    for measure in _measures:
        filename = f'GLC {measure} daily data current month.xlsx'
        stream = archive.open(filename, 'r')
        wb = openpyxl.load_workbook(stream)

        # Spot curve, long end
        sh = wb['4. spot curve']

        date, df = read(sh, f'{measure}_Spot')

        dfs.append(df)

    df = pd.concat(dfs, axis=1)

    df.interpolate(method='cubicspline', axis=0, limit_direction='both', inplace=True)

    df.to_csv(_filename, float_format='{:.6f}'.format)

    if False:
        df.plot()
        import matplotlib.pyplot as plt
        plt.show()


def yield_curves():
    if not os.path.exists(_filename):
        load()
    df = pd.read_csv(_filename, header=0, index_col=0)
    for _measure in _measures:
        df[f'{_measure}_Spot'] = df[f'{_measure}_Spot'].multiply(.01)
    return df


if __name__ == '__main__':
    load()